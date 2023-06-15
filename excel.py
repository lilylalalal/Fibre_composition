from openpyxl import Workbook
##open an excel file
workbk = Workbook()
wsheet1 = workbk.active
wsheet2 = workbk.create_sheet("Testing")
print(workbk.sheetnames)
wsheet1['C9'] = 'hello world'
addingvalue = wsheet2.cell(row=4, column=2, value=10)
workbk.save('testing.xlsx')

##
================

from openpyxl import load_workbook
##read an excel from the file
wb = load_workbook(filename = "AW23 MASTER LINESHEET.xlsx")
sheet_range = wb['AW23 RTW']
print(sheet_range['AG38'].value)
##